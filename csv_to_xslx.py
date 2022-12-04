import csv
from os import listdir, getcwd
from os.path import isfile, join
from openpyxl import Workbook

# mypath = getcwd()
current_dir = getcwd()
transfer_dir = "C:\\Users\\PB088\\Desktop\\Cell counts\\"

csv_files = [f for f in listdir(mypath) if isfile(join(mypath, f)) and f[-3:] == "csv"]

print(csv_files)

for csv_file in csv_files:
    wb = Workbook()
    ws = wb.active 
    with open(mypath + csv_file) as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader, start=1):
            if len(row) == 11:
                for j in range(11):
                    ws.cell(column=j+1, row=i).value = row[j]
                    # print(ws.cell(column=j+1, row=i).value)

        wb.save(filename=str(mypath + csv_file[0:-3]) + "xlsx")


